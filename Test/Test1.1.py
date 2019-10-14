from selenium import webdriver
import time
import os
import openpyxl
from openpyxl import Workbook
#options = webdriver.ChromeOptions()
#options.add_argumnet("__incognito")
driver = webdriver.Chrome("E:/chromedriver/chromedriver.exe") #chrome_options=options
#driver.maximize_window()
driver.get("https://devicemon.azurewebsites.net")
driver.find_element_by_xpath("//*[@id='i0116']").send_keys("test.ncadminuser@ha-efbops.com")
driver.implicitly_wait(50)
driver.find_element_by_id("idSIButton9").click()
#driver.implicitly_wait(40)
driver.find_element_by_xpath("//*[@id='i0118']").send_keys("Nathcorp!1")
time.sleep(4)
driver.find_element_by_id("idSIButton9").click()
driver.find_element_by_id("idSIButton9").click()
#driver.implicitly_wait(30)
driver.find_element_by_id("TableSearch").send_keys("014269671953")
driver.find_element_by_xpath("//a[@id='App']").click()
time.sleep(2)

row_count = len(driver.find_elements_by_xpath("//*[@id='PidAppInfoDetails']/tr "))
col_count = len(driver.find_elements_by_xpath("//*[@id='PidAppInfoDetails']/tr[1]/td"))

print(row_count)
print(col_count)
first_part = "//*[@id='PidAppInfoDetails']/tr[ "
second_part = "]/td["
Third_part = "]"
for n in range(1,row_count+1):
    for m in range(1, col_count + 1):
        final_path = first_part+str(n)+second_part+str(m)+Third_part
        table_data= driver.find_element_by_xpath(final_path).text
        fname= 'applist.xlsx'
        if(os.path.exists(fname)):
            Workbook =openpyxl.load_workbook(fname)
            Worksheet = Workbook.get_sheet_by_name('Sheet')
        else:
            Workbook =Workbook()
            Worksheet =Workbook.active
        Worksheet.cell(row=n,column=m).value = table_data
        Workbook.save(fname)
        print(table_data,end =" ")
        print("")
        #
driver.find_element_by_xpath("//*[@id='PidAppInfoPopup']/div/div/div[3]/button").click()
driver.find_element_by_xpath("//*[@id='014269671953]").click()



driver.find_element_by_xpath("//*[@id='UserLoginInfoPopup]/div/div/div[3]/button").click()

'''
def app_list(url):
    thepage= urllib.request.urlopen(url) //*[@id="UserLoginInfoPopup"]/div/div/div[3]/button
    soupdata =BeautifulSoup(thepage,"csv")
    return soupdata
soup= app_list("https://devicemon.azurewebsites.net")
for record in soup.findAll('tr'):
    for data in record.findAll('td')
   //*[@id="PidAppInfoDetails"]/tr[1]     
'''
'''
posts= driver.find_element_by_id("PidAppInfoDetails")
for post in posts:
    print(post.text)
'''
'''driver.close()'''